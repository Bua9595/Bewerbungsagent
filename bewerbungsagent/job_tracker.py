from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .job_state import (
    STATUS_APPLIED,
    STATUS_CLOSED,
    STATUS_IGNORED,
    now_iso,
    parse_ts,
)

TRACKER_HEADERS = [
    "job_uid",
    "status",
    "applied_at",
    "erledigt",
    "aktion",
    "title",
    "company",
    "location",
    "source",
    "link",
    "first_seen_at",
    "last_seen_at",
    "last_sent_at",
    "score",
    "match",
    "notes",
]

MANUAL_COLUMNS = {"erledigt", "aktion", "notes"}

CHECKBOX_EMPTY = chr(0x2610)
CHECKBOX_DONE = chr(0x2611)
CHECKBOX_VALUES = (CHECKBOX_EMPTY, CHECKBOX_DONE)

TRUTHY = {"1", "true", "t", "yes", "y", "ja", "j", "x", CHECKBOX_DONE}
APPLIED_ACTIONS = {"applied", "apply", "done", "sent", "bewerbung", "gesendet"}
IGNORED_ACTIONS = {"ignored", "ignore", "skip", "no", "nein"}


def get_tracker_path() -> Path:
    return Path(os.getenv("JOB_TRACKER_FILE", "generated/job_tracker.xlsx"))


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() == ".xlsx"


def _normalize_erledigt(value: Any) -> str:
    raw = _clean(value)
    if not raw:
        return CHECKBOX_EMPTY
    lowered = raw.lower()
    if raw in CHECKBOX_VALUES:
        return raw
    if lowered in TRUTHY:
        return CHECKBOX_DONE
    if lowered in {"0", "false", "no", "nein"}:
        return CHECKBOX_EMPTY
    return raw


def load_tracker(path: Path) -> Dict[str, Dict[str, Any]]:
    if not path.exists():
        if _is_xlsx(path):
            legacy = path.with_suffix(".csv")
            if legacy.exists():
                return _load_tracker_csv(legacy)
        return {}
    if _is_xlsx(path):
        return _load_tracker_xlsx(path)
    return _load_tracker_csv(path)


def _load_tracker_csv(path: Path) -> Dict[str, Dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "job_uid" not in reader.fieldnames:
            return {}
        rows: Dict[str, Dict[str, Any]] = {}
        for row in reader:
            uid = _clean(row.get("job_uid"))
            if not uid:
                continue
            rows[uid] = row
        return rows


def _load_tracker_xlsx(path: Path) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        wb.close()
        return {}
    header_list = [str(h).strip() if h is not None else "" for h in headers]
    if "job_uid" not in header_list:
        wb.close()
        return {}
    rows: Dict[str, Dict[str, Any]] = {}
    for values in iterator:
        row: Dict[str, Any] = {}
        for idx, header in enumerate(header_list):
            if not header:
                continue
            val = values[idx] if idx < len(values) else ""
            row[header] = "" if val is None else str(val).strip()
        uid = _clean(row.get("job_uid"))
        if not uid:
            continue
        rows[uid] = row
    wb.close()
    return rows


def apply_tracker_marks(
    state: Dict[str, Dict[str, Any]],
    tracker_rows: Dict[str, Dict[str, Any]],
) -> int:
    updates = 0
    stamp = now_iso()
    for uid, row in tracker_rows.items():
        record = state.get(uid)
        if not record:
            continue
        action = _clean(row.get("aktion")).lower()
        done = _normalize_erledigt(row.get("erledigt"))
        desired = ""
        if action in APPLIED_ACTIONS:
            desired = STATUS_APPLIED
        elif action in IGNORED_ACTIONS:
            desired = STATUS_IGNORED
        elif done == CHECKBOX_DONE:
            desired = STATUS_APPLIED
        if desired and record.get("status") != desired:
            record["status"] = desired
            if desired == STATUS_APPLIED:
                record["applied_at"] = stamp
            else:
                record.pop("applied_at", None)
            updates += 1
    return updates


def _sort_key(row: Dict[str, Any]) -> float:
    last_seen = parse_ts(_clean(row.get("last_seen_at")))
    return last_seen.timestamp() if last_seen else 0.0


def build_tracker_rows(
    state: Dict[str, Dict[str, Any]],
    existing_rows: Dict[str, Dict[str, Any]] | None = None,
    include_closed: bool = False,
) -> list[Dict[str, Any]]:
    existing_rows = existing_rows or {}
    rows: list[Dict[str, Any]] = []
    for uid, record in state.items():
        status = record.get("status") or ""
        if status == STATUS_CLOSED and not include_closed:
            continue
        row = {k: "" for k in TRACKER_HEADERS}
        row.update(
            {
                "job_uid": uid,
                "status": status,
                "applied_at": record.get("applied_at") or "",
                "erledigt": CHECKBOX_EMPTY,
                "title": record.get("title") or "",
                "company": record.get("company") or "",
                "location": record.get("location") or "",
                "source": record.get("source") or "",
                "link": record.get("link") or record.get("canonical_url") or "",
                "first_seen_at": record.get("first_seen_at") or "",
                "last_seen_at": record.get("last_seen_at") or "",
                "last_sent_at": record.get("last_sent_at") or "",
                "score": record.get("score") or "",
                "match": record.get("match") or "",
            }
        )
        existing = existing_rows.get(uid, {})
        for col in MANUAL_COLUMNS:
            if not _clean(existing.get(col)):
                continue
            if col == "erledigt":
                row[col] = _normalize_erledigt(existing.get(col))
            else:
                row[col] = existing.get(col)
        if status in (STATUS_APPLIED, STATUS_IGNORED):
            row["erledigt"] = CHECKBOX_DONE
            if status == STATUS_APPLIED and not row["aktion"]:
                row["aktion"] = "applied"
            if status == STATUS_IGNORED and not row["aktion"]:
                row["aktion"] = "ignored"
        rows.append(row)

    rows.sort(key=_sort_key, reverse=True)
    return rows


def write_tracker(
    state: Dict[str, Dict[str, Any]],
    path: Path,
    existing_rows: Dict[str, Dict[str, Any]] | None = None,
    include_closed: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = build_tracker_rows(state, existing_rows, include_closed)
    if _is_xlsx(path):
        _write_tracker_xlsx(path, rows)
    else:
        _write_tracker_csv(path, rows)


def _write_tracker_csv(path: Path, rows: list[Dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TRACKER_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def _write_tracker_xlsx(path: Path, rows: list[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "job_tracker"
    ws.append(TRACKER_HEADERS)
    for row in rows:
        ws.append([row.get(col, "") for col in TRACKER_HEADERS])

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(TRACKER_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"

    erledigt_idx = TRACKER_HEADERS.index("erledigt") + 1
    col_letter = get_column_letter(erledigt_idx)
    dv = DataValidation(
        type="list",
        formula1=f'"{CHECKBOX_EMPTY},{CHECKBOX_DONE}"',
        allow_blank=False,
    )
    dv.error = f"Bitte nur {CHECKBOX_EMPTY} oder {CHECKBOX_DONE} waehlen."
    dv.errorTitle = "Ungueltiger Wert"
    ws.add_data_validation(dv)
    if rows:
        dv.add(f"{col_letter}2:{col_letter}{len(rows) + 1}")
    wb.save(path)
